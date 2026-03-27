"""
Dashboard Flask para seguimiento de entregas de estudiantes.
Lee datos de SQLite de NocoDB en modo read-only.
"""

import os
import sqlite3
import logging
from datetime import datetime
from typing import Optional, Dict, List, Any, Tuple
from flask import Flask, render_template, request, jsonify, send_file, Response
import io
from functools import wraps

# Configuración de logs
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# === CONFIGURACIÓN ===
DB_PATH = os.environ.get('NOCODB_DB_PATH', '/data/noco.db')
FILTRAR_ACTIVAS = os.environ.get('COHORTE_FILTRAR_ACTIVAS', 'true').lower() == 'true'

# Credenciales de administrador (docente) para rutas privadas
ADMIN_USER = os.environ.get('ADMIN_USER', 'docente')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'entregas2026')

if not os.environ.get('ADMIN_PASSWORD'):
    logger.warning("[SEGURIDAD] ADMIN_PASSWORD no configurada. Usando password por defecto. "
                   "Configura ADMIN_PASSWORD en .env para produccion.")

# === AUTENTICACIÓN ===
def check_auth(username, password):
    """Verifica si las credenciales coinciden con las del entorno."""
    return username == ADMIN_USER and password == ADMIN_PASSWORD

def authenticate():
    """Envía un 401 que solicita al navegador mostrar la ventana de login."""
    return Response(
    'No autorizado. Ingrese sus credenciales de docente.\n', 401,
    {'WWW-Authenticate': 'Basic realm="Acceso Docentes"'})

def requires_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization
        if not auth or not check_auth(auth.username, auth.password):
            return authenticate()
        return f(*args, **kwargs)
    return decorated

# Nombres visibles de tablas que esperamos encontrar
# El prefijo cambia según el ID del proyecto en NocoDB (v2 suele ser nc_ID___)
TABLE_TITLES = {
    'Cohortes': None,
    'Actividades': None,
    'Estudiantes': None,
    'Entregas': None
}

# Estado global de validación (se llena al arrancar)
_schema_warnings: List[str] = []
_schema_fixed: bool = False


# === CAPA DB ===

def get_db(writable: bool = False):
    """
    Conexión SQLite con timeout.
    Por defecto abre en modo read-only. Usar writable=True solo para auto-fix.
    """
    try:
        if os.path.exists(DB_PATH) and not writable:
            uri = f"file:{DB_PATH}?mode=ro"
            conn = sqlite3.connect(uri, uri=True, timeout=20.0)
        else:
            conn = sqlite3.connect(DB_PATH, timeout=20.0)
        conn.row_factory = sqlite3.Row
        return conn
    except Exception as e:
        logger.error(f"Error conectando a la base de datos: {e}")
        conn = sqlite3.connect(DB_PATH, timeout=20.0)
        conn.row_factory = sqlite3.Row
        return conn


def discover_schema(conn: sqlite3.Connection) -> Dict[str, Dict[str, Any]]:
    """
    Descubre el schema real de NocoDB leyendo las tablas de metadatos de forma flexible.
    """
    schema = {}
    cursor = conn.cursor()

    try:
        # Verificar tablas de metadatos (NocoDB v2)
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='nc_models_v2'")
        if not cursor.fetchone():
            return {}

        # 1. Leer modelos (Tablas)
        cursor.execute("SELECT id, table_name, title FROM nc_models_v2")
        models_data = cursor.fetchall()
        
        # Mapeo de Título -> Datos de tabla
        # Usamos búsqueda flexible (insensible a mayúsculas y plurales comunes)
        table_targets = {
            'Cohortes': ['Cohortes', 'Cohorte', 'Cursos'],
            'Actividades': ['Actividades', 'Actividad', 'Tareas'],
            'Estudiantes': ['Estudiantes', 'Estudiante', 'Alumnos'],
            'Entregas': ['Entregas', 'Entrega', 'TPs']
        }

        found_models = {} # title_normalizado -> {id, table_name, title_original}
        for row in models_data:
            t_orig = row['title']
            for norm, aliases in table_targets.items():
                if any(a.lower() == t_orig.lower() for a in aliases):
                    found_models[norm] = {'id': row['id'], 'table_name': row['table_name'], 'title': t_orig}
                    break

        # 2. Leer todas las columnas registradas en NocoDB
        cursor.execute("SELECT fk_model_id, column_name, title, uidt FROM nc_columns_v2")
        columns_data = cursor.fetchall()
        
        cols_by_model = {}
        for row in columns_data:
            mid = row['fk_model_id']
            if mid not in cols_by_model:
                cols_by_model[mid] = []
            cols_by_model[mid].append(row)

        # 3. Construir esquema base
        for norm_title, m_info in found_models.items():
            mid = m_info['id']
            schema[norm_title] = {
                'id': mid,
                'table_name': m_info['table_name'],
                'columns': {row['title']: row['column_name'] for row in cols_by_model.get(mid, []) if row['column_name']},
                'links': [row for row in cols_by_model.get(mid, []) if row['uidt'] == 'LinkToAnotherRecord']
            }

        # 4. TRUCO DE MAGIA: Si no hay FK directa, buscar si el padre tiene la FK (Relación Inversa)
        # O si hay una columna física 'nc_..._id' que no esté mapeada en NocoDB
        for child, parent in [('Estudiantes', 'Cohortes'), ('Actividades', 'Cohortes'), 
                             ('Entregas', 'Estudiantes'), ('Entregas', 'Actividades')]:
            if child in schema and parent in schema:
                # ¿Tiene el hijo una columna que apunte al padre?
                fk_found = False
                child_cols = schema[child]['columns']
                
                # Buscar por nombre físico en la tabla
                cursor.execute(f"PRAGMA table_info({schema[child]['table_name']})")
                physical_cols = [r['name'] for r in cursor.fetchall()]
                
                # Buscar cualquier columna que parezca un ID del padre
                parent_phys_base = schema[parent]['table_name']
                possible_fks = [c for c in physical_cols if parent_phys_base in c or parent.lower()[:-2] in c.lower()]
                
                if possible_fks:
                    schema[child]['columns'][f'fk_{parent}'] = possible_fks[0]
                    fk_found = True
                
                if not fk_found:
                    # Buscar en el PADRE si existe un ID que apunte al HIJO (Relación invertida accidental)
                    cursor.execute(f"PRAGMA table_info({schema[parent]['table_name']})")
                    parent_phys_cols = [r['name'] for r in cursor.fetchall()]
                    child_phys_base = schema[child]['table_name']
                    
                    rev_fks = [c for c in parent_phys_cols if child_phys_base in c or child.lower()[:-2] in c.lower()]
                    if rev_fks:
                        schema[child]['columns'][f'rev_fk_{parent}'] = rev_fks[0]
                        logger.info(f"Detectada relación inversa: {parent}.{rev_fks[0]} apunta a {child}")

        return schema

    except Exception as e:
        logger.error(f"Error crítico en discover_schema: {e}")
        return {}


# === VALIDACIÓN Y AUTO-FIX DEL ESQUEMA ===

# Relaciones esperadas: (tabla_hija, columna_fk_título, tabla_padre)
EXPECTED_RELATIONS = [
    ('Estudiantes', 'Cohorte', 'Cohortes'),
    ('Actividades', 'Cohorte', 'Cohortes'),
    ('Entregas', 'Estudiante', 'Estudiantes'),
    ('Entregas', 'Actividad', 'Actividades'),
]


def validate_schema(conn: sqlite3.Connection, schema: Dict) -> Dict[str, Any]:
    """
    Valida el esquema de NocoDB detectando:
    - Tablas faltantes
    - Relaciones M2M (tablas intermedias m2m_) en vez de FK directas
    - Columnas FK faltantes en tablas hijas
    Retorna: {"ok": bool, "warnings": [...], "m2m_tables": [...], "missing_tables": [...]}
    """
    result = {"ok": True, "warnings": [], "m2m_tables": [], "missing_tables": []}
    cursor = conn.cursor()

    # 1. Verificar que las 4 tablas existen
    for title in ['Cohortes', 'Actividades', 'Estudiantes', 'Entregas']:
        table_name = get_table_name(schema, title)
        if not table_name:
            result['missing_tables'].append(title)
            result['warnings'].append(f"Tabla '{title}' no encontrada en el esquema")
            result['ok'] = False
            continue
        try:
            cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
        except sqlite3.OperationalError:
            result['missing_tables'].append(title)
            result['warnings'].append(f"Tabla '{title}' ({table_name}) no existe en la DB")
            result['ok'] = False

    # 2. Detectar tablas M2M (señal de relaciones incorrectas)
    try:
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name LIKE '%m2m%'")
        m2m_tables = [row['name'] for row in cursor.fetchall()]
        if m2m_tables:
            result['m2m_tables'] = m2m_tables
            result['warnings'].append(
                f"Detectadas tablas M2M: {', '.join(m2m_tables)}. "
                "Esto indica relaciones Many-to-Many en vez de FK directas (BelongsTo). "
                "Se intentará auto-fix."
            )
            result['ok'] = False
    except Exception as e:
        logger.warning(f"Error buscando tablas M2M: {e}")

    # 3. Verificar que las FK directas existen (Soporte Smart)
    for child_title, fk_title, parent_title in EXPECTED_RELATIONS:
        if child_title in result['missing_tables'] or parent_title in result['missing_tables']:
            continue
        
        fk_col = get_column_name(schema, child_title, fk_title)
        if not fk_col:
            # Si no hay FK, ¿hay relación inversa?
            if f"rev_fk_{parent_title}" in schema.get(child_title, {}).get('columns', {}):
                continue # Todo ok, es una relación inversa detectada
                
            result['warnings'].append(
                f"Relación faltante: '{child_title}' -> '{parent_title}'"
            )
            result['ok'] = False

    return result


def attempt_schema_fix(conn_rw: sqlite3.Connection, schema: Dict,
                       validation: Dict[str, Any]) -> bool:
    """
    Intenta corregir relaciones M2M migrando datos a FK directas.
    Usa una conexión de escritura.
    
    Estrategia:
    1. Para cada tabla M2M detectada, identifica las tablas que conecta.
    2. Crea una columna FK directa en la tabla hija si no existe.
    3. Migra los datos de la tabla M2M a la nueva columna FK.
    4. No elimina las tablas M2M (NocoDB las gestiona).

    Retorna True si el fix fue exitoso.
    """
    if not validation['m2m_tables']:
        return False

    cursor = conn_rw.cursor()
    fixed_something = False

    for m2m_table in validation['m2m_tables']:
        try:
            logger.info(f"[AUTO-FIX] Analizando tabla M2M: {m2m_table}")

            # Obtener columnas de la tabla M2M
            cursor.execute(f"PRAGMA table_info({m2m_table})")
            m2m_cols = [row['name'] for row in cursor.fetchall()]
            logger.info(f"[AUTO-FIX] Columnas en {m2m_table}: {m2m_cols}")

            # Las tablas M2M de NocoDB tienen columnas como:
            # id, <tabla1>_id, <tabla2>_id
            fk_cols = [c for c in m2m_cols if c != 'id' and c.endswith('_id')]
            if len(fk_cols) != 2:
                # Intentar buscar por otro patrón
                fk_cols = [c for c in m2m_cols if c != 'id']

            if len(fk_cols) < 2:
                logger.warning(f"[AUTO-FIX] No se encontraron 2 FK en {m2m_table}: {fk_cols}")
                continue

            logger.info(f"[AUTO-FIX] FK detectadas en M2M: {fk_cols}")

            # Determinar cuál es la tabla hija y cuál la padre
            # basándose en las relaciones esperadas
            for child_title, fk_title, parent_title in EXPECTED_RELATIONS:
                child_table = get_table_name(schema, child_title)
                parent_table = get_table_name(schema, parent_title)
                if not child_table or not parent_table:
                    continue

                # Verificar si esta tabla M2M conecta a estas dos tablas
                child_match = any(child_title.lower() in c.lower() for c in fk_cols)
                parent_match = any(parent_title.lower() in c.lower() for c in fk_cols)

                if not (child_match and parent_match):
                    continue

                # Identificar las columnas FK en la M2M
                child_fk_in_m2m = next((c for c in fk_cols if child_title.lower() in c.lower()), None)
                parent_fk_in_m2m = next((c for c in fk_cols if parent_title.lower() in c.lower()), None)

                if not child_fk_in_m2m or not parent_fk_in_m2m:
                    continue

                # Nombre de la nueva columna FK directa
                expected_fk_col = get_column_name(schema, child_title, fk_title)
                if not expected_fk_col:
                    expected_fk_col = f"nc_fk_{parent_title.lower()}_id"

                # Verificar si la columna FK ya existe en la tabla hija
                cursor.execute(f"PRAGMA table_info({child_table})")
                existing_cols = [row['name'] for row in cursor.fetchall()]

                if expected_fk_col not in existing_cols:
                    logger.info(
                        f"[AUTO-FIX] Creando columna {expected_fk_col} en {child_table}"
                    )
                    cursor.execute(
                        f"ALTER TABLE {child_table} ADD COLUMN {expected_fk_col} INTEGER"
                    )
                    fixed_something = True

                # Migrar datos de la tabla M2M a la FK directa
                cursor.execute(f"SELECT {child_fk_in_m2m}, {parent_fk_in_m2m} FROM {m2m_table}")
                m2m_data = cursor.fetchall()
                logger.info(
                    f"[AUTO-FIX] Migrando {len(m2m_data)} relaciones de {m2m_table} "
                    f"a {child_table}.{expected_fk_col}"
                )

                for row in m2m_data:
                    child_id = row[child_fk_in_m2m]
                    parent_id = row[parent_fk_in_m2m]
                    cursor.execute(
                        f"UPDATE {child_table} SET {expected_fk_col} = ? WHERE id = ?",
                        (parent_id, child_id)
                    )
                    fixed_something = True

                logger.info(
                    f"[AUTO-FIX] ✓ Migración completada: {m2m_table} → "
                    f"{child_table}.{expected_fk_col}"
                )

        except Exception as e:
            logger.error(f"[AUTO-FIX] Error procesando {m2m_table}: {e}")
            continue

    if fixed_something:
        conn_rw.commit()
        logger.info("[AUTO-FIX] ✓ Todos los cambios committeados")

    return fixed_something


def run_startup_validation():
    """
    Ejecuta validación y auto-fix del esquema al arrancar la app.
    Se ejecuta una sola vez. Llena las variables globales de estado.
    """
    global _schema_warnings, _schema_fixed

    if not os.path.exists(DB_PATH):
        _schema_warnings = [f"Base de datos no encontrada en {DB_PATH}"]
        logger.warning(f"[STARTUP] DB no encontrada: {DB_PATH}")
        return

    try:
        conn = get_db()
        schema = discover_schema(conn)
        validation = validate_schema(conn, schema)
        conn.close()

        if validation['ok']:
            logger.info("[STARTUP] ✓ Esquema validado correctamente")
            return

        # Hay problemas — intentar auto-fix
        _schema_warnings = validation['warnings']
        for w in _schema_warnings:
            logger.warning(f"[STARTUP] {w}")

        if validation['m2m_tables']:
            logger.info("[STARTUP] Intentando auto-fix de relaciones M2M...")
            try:
                conn_rw = get_db(writable=True)
                schema = discover_schema(conn_rw)
                _schema_fixed = attempt_schema_fix(conn_rw, schema, validation)
                conn_rw.close()

                if _schema_fixed:
                    logger.info("[STARTUP] ✓ Auto-fix exitoso")
                    # Re-validar
                    conn = get_db()
                    schema = discover_schema(conn)
                    revalidation = validate_schema(conn, schema)
                    conn.close()
                    if revalidation['ok']:
                        _schema_warnings = []
                        logger.info("[STARTUP] ✓ Re-validación exitosa")
                    else:
                        _schema_warnings = revalidation['warnings']
                else:
                    logger.warning("[STARTUP] Auto-fix no pudo corregir los problemas")
            except Exception as e:
                logger.error(f"[STARTUP] Error en auto-fix: {e}")
    except Exception as e:
        _schema_warnings = [f"Error durante validación: {e}"]
        logger.error(f"[STARTUP] Error general: {e}")


def get_column_name(schema: Dict, table_title: str, column_title: str) -> Optional[str]:
    """Obtiene el nombre físico de una columna con soporte para alias comunes."""
    if table_title not in schema:
        return None
    
    cols = schema[table_title]['columns']
    
    # Intento 1: Nombre exacto (case insensitive)
    for c_title, c_phys in cols.items():
        if c_title.lower() == column_title.lower():
            return c_phys
    
    # Intento 2: Búsqueda por sub-string del título
    for c_title, c_phys in cols.items():
        if column_title.lower() in c_title.lower():
            return c_phys

    # Intento 3: Si es una FK pre-calculada en discover_schema
    if f"fk_{column_title}" in cols:
        return cols[f"fk_{column_title}"]
    if f"rev_fk_{column_title}" in cols:
        return cols[f"rev_fk_{column_title}"]

    # Intento 4: Alias predefinidos
    aliases = {
        'Fecha Límite': ['Fecha', 'Deadline', 'Fecha_Limite', 'Entrega', 'Limit'],
        'URL Guía': ['URL_Guia', 'URL_Gu_a', 'Guia', 'Link', 'Consigna'],
        'Activa': ['Activo', 'Status', 'Active'],
        'Nombre': ['Title', 'Name', 'Nombres'],
        'Apellido': ['Lastname', 'Surname', 'Apellidos'],
        'URL Entrega': ['URL', 'Link', 'Entrega', 'Repo'],
        'Estudiante': ['Estudiante_id', 'Alumno', 'User'],
        'Actividad': ['Actividad_id', 'Tarea', 'Task'],
        'Cohorte': ['Cohorte_id', 'Curso', 'Group']
    }
    
    if column_title in aliases:
        target_aliases = [a.lower() for a in aliases[column_title]]
        for c_title, c_phys in cols.items():
            if c_title.lower() in target_aliases:
                return c_phys
                
    return None


def get_table_name(schema: Dict, table_title: str) -> Optional[str]:
    """Obtiene el nombre físico de una tabla."""
    if table_title not in schema:
        return None
    return schema[table_title]['table_name']


# === CAPA DATOS ===

def get_cohortes(conn: sqlite3.Connection, schema: Dict) -> List[Dict]:
    """Lista todas las cohortes activas."""
    table = get_table_name(schema, 'Cohortes')
    if not table:
        return []

    col_id = 'id'
    col_nombre = get_column_name(schema, 'Cohortes', 'Nombre') or 'Nombre'
    col_activa = get_column_name(schema, 'Cohortes', 'Activa') or 'Activa'

    cursor = conn.cursor()
    
    try:
        if FILTRAR_ACTIVAS:
            query = f"SELECT {col_id} as id, {col_nombre} as nombre FROM {table} WHERE {col_activa} = 1 ORDER BY {col_nombre}"
        else:
            query = f"SELECT {col_id} as id, {col_nombre} as nombre FROM {table} ORDER BY {col_nombre}"
        
        cursor.execute(query)
        return [dict(row) for row in cursor.fetchall()]
    except sqlite3.OperationalError as e:
        logger.error(f"Error consultando cohortes: {e}")
        return []


def get_matrix_data(conn: sqlite3.Connection, schema: Dict, cohorte_id: int) -> Tuple[List, List, Dict]:
    """Obtiene datos para la matriz estudiante x actividad."""
    table_est = get_table_name(schema, 'Estudiantes')
    table_act = get_table_name(schema, 'Actividades')
    table_ent = get_table_name(schema, 'Entregas')

    if not all([table_est, table_act, table_ent]):
        return [], [], {}

    cursor = conn.cursor()

    # Resolver nombres de columnas
    col_est_id = 'id'
    col_est_nombre = get_column_name(schema, 'Estudiantes', 'Nombre') or 'Nombre'
    col_est_apellido = get_column_name(schema, 'Estudiantes', 'Apellido') or 'Apellido'
    col_est_email = get_column_name(schema, 'Estudiantes', 'Email') or 'Email'
    col_est_github = get_column_name(schema, 'Estudiantes', 'GitHub') or 'GitHub'
    col_est_cohorte = get_column_name(schema, 'Estudiantes', 'Cohorte') or 'nc_fk_cohorte_id'

    col_act_id = 'id'
    col_act_nombre = get_column_name(schema, 'Actividades', 'Nombre') or 'Nombre'
    col_act_orden = get_column_name(schema, 'Actividades', 'Orden') or 'Orden'
    col_act_fecha_limite = get_column_name(schema, 'Actividades', 'Fecha Límite') or 'Fecha'
    col_act_peso = get_column_name(schema, 'Actividades', 'Peso') or 'Peso'
    col_act_cohorte = get_column_name(schema, 'Actividades', 'Cohorte') or 'nc_fk_cohorte_id'

    col_ent_estado = get_column_name(schema, 'Entregas', 'Estado') or 'Estado'
    col_ent_url = get_column_name(schema, 'Entregas', 'URL Entrega') or 'URL_Entrega'
    col_ent_estudiante = get_column_name(schema, 'Entregas', 'Estudiante') or 'nc_fk_estudiante_id'
    col_ent_actividad = get_column_name(schema, 'Entregas', 'Actividad') or 'nc_fk_actividad_id'

    try:
        # 1. Estudiantes (Soporte para relación normal o inversa)
        if 'rev_fk_Cohortes' in schema['Estudiantes']['columns']:
            rev_fk = schema['Estudiantes']['columns']['rev_fk_Cohortes']
            table_coh = get_table_name(schema, 'Cohortes')
            query_est = f"""
                SELECT s.{col_est_id} as id, s.{col_est_nombre} as nombre, s.{col_est_apellido} as apellido, 
                       s.{col_est_email} as email, s.{col_est_github} as github 
                FROM {table_est} s
                JOIN {table_coh} c ON c.{rev_fk} = s.{col_est_id}
                WHERE c.id = ?
                ORDER BY s.{col_est_apellido}
            """
            cursor.execute(query_est, (cohorte_id,))
        else:
            cursor.execute(f"SELECT {col_est_id} as id, {col_est_nombre} as nombre, {col_est_apellido} as apellido, {col_est_email} as email, {col_est_github} as github FROM {table_est} WHERE {col_est_cohorte} = ? ORDER BY {col_est_apellido}", (cohorte_id,))
        students = [dict(row) for row in cursor.fetchall()]

        # 2. Actividades (Soporte para relación normal o inversa)
        if 'rev_fk_Cohortes' in schema['Actividades']['columns']:
            rev_fk = schema['Actividades']['columns']['rev_fk_Cohortes']
            table_coh = get_table_name(schema, 'Cohortes')
            query_act = f"""
                SELECT a.{col_act_id} as id, a.{col_act_nombre} as nombre, a.{col_act_fecha_limite} as fecha_limite, a.{col_act_peso} as peso 
                FROM {table_act} a
                JOIN {table_coh} c ON c.{rev_fk} = a.{col_act_id}
                WHERE c.id = ?
                ORDER BY a.{col_act_orden}
            """
            cursor.execute(query_act, (cohorte_id,))
        else:
            cursor.execute(f"SELECT {col_act_id} as id, {col_act_nombre} as nombre, {col_act_fecha_limite} as fecha_limite, {col_act_peso} as peso FROM {table_act} WHERE {col_act_cohorte} = ? ORDER BY {col_act_orden}", (cohorte_id,))
        activities = [dict(row) for row in cursor.fetchall()]

        # 3. Entregas (Búsqueda compleja con soporte para relaciones invertidas)
        student_ids = [s['id'] for s in students]
        activity_ids = [a['id'] for a in activities]
        cells = {}

        if student_ids and activity_ids:
            # Construcción dinámica de la query de entregas
            placeholders_s = ','.join('?' * len(student_ids))
            placeholders_a = ','.join('?' * len(activity_ids))
            
            # Caso base (Relación normal: FK en Entregas)
            query_ent = f"SELECT {col_ent_estudiante} as estudiante_id, {col_ent_actividad} as actividad_id, {col_ent_estado} as estado, {col_ent_url} as url, created_at as fecha_entrega FROM {table_ent} WHERE {col_ent_estudiante} IN ({placeholders_s}) AND {col_ent_actividad} IN ({placeholders_a})"
            
            # Ajuste de query si hay relaciones invertidas
            # Si Estudiante tiene la FK de Entrega
            if 'rev_fk_Estudiantes' in schema['Entregas']['columns']:
                rev_fk_est = schema['Entregas']['columns']['rev_fk_Estudiantes']
                query_ent = query_ent.replace(f"WHERE {col_ent_estudiante}", f"JOIN {table_est} s_fk ON s_fk.{rev_fk_est} = {table_ent}.id WHERE s_fk.id")
                # Necesitamos estudiante_id en el select
                query_ent = query_ent.replace(f"SELECT {col_ent_estudiante} as estudiante_id", f"SELECT s_fk.id as estudiante_id")

            # Si Actividad tiene la FK de Entrega
            if 'rev_fk_Actividades' in schema['Entregas']['columns']:
                rev_fk_act = schema['Entregas']['columns']['rev_fk_Actividades']
                query_ent = query_ent.replace(f"AND {col_ent_actividad}", f"JOIN {table_act} a_fk ON a_fk.{rev_fk_act} = {table_ent}.id WHERE a_fk.id")
                query_ent = query_ent.replace(f"{col_ent_actividad} as actividad_id", f"a_fk.id as actividad_id")

            cursor.execute(query_ent, student_ids + activity_ids)
            for row in cursor.fetchall():
                cells[(row['estudiante_id'], row['actividad_id'])] = dict(row)

        return students, activities, cells
    except Exception as e:
        logger.error(f"Error en get_matrix_data: {e}")
        return [], [], {}


def parse_date(date_str: Optional[str]) -> Optional[datetime]:
    if not date_str: return None
    formats = ['%Y-%m-%dT%H:%M:%S.%f%z', '%Y-%m-%dT%H:%M:%S%z', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d']
    for fmt in formats:
        try: return datetime.strptime(date_str, fmt)
        except: continue
    return None


def compute_cell(entrega: Optional[Dict], fecha_limite_str: Optional[str], now: datetime) -> Dict:
    fecha_limite = parse_date(fecha_limite_str)
    COLOR_VERDE, COLOR_AMARILLO, COLOR_ROJO, COLOR_GRIS = '#22c55e', '#eab308', '#ef4444', '#d1d5db'
    
    if not entrega:
        if fecha_limite and fecha_limite < now:
            return {'color': COLOR_ROJO, 'status': 'No entregado', 'label': '✗', 'tooltip': 'No entregado'}
        return {'color': COLOR_GRIS, 'status': 'Pendiente', 'label': '−', 'tooltip': 'Pendiente'}

    estado = entrega.get('estado', 'Entregado')
    fecha_entrega = parse_date(entrega.get('fecha_entrega'))
    esta_tarde = (fecha_entrega > fecha_limite) if (fecha_entrega and fecha_limite) else False
    
    tooltip = f"{estado}"
    if fecha_entrega: tooltip += f" ({fecha_entrega.strftime('%d/%m')})"

    if estado == 'Corregido': return {'color': COLOR_VERDE, 'status': 'Corregido', 'label': '✓', 'tooltip': tooltip}
    if estado == 'Rehacer': return {'color': COLOR_AMARILLO, 'status': 'Rehacer', 'label': '↻', 'tooltip': tooltip}
    if esta_tarde: return {'color': COLOR_AMARILLO, 'status': 'Tarde', 'label': '⚠', 'tooltip': tooltip}
    return {'color': COLOR_VERDE, 'status': 'Entregado', 'label': '✓', 'tooltip': tooltip}


@app.route('/')
def index():
    conn = get_db()
    try:
        schema = discover_schema(conn)
        cohortes = get_cohortes(conn, schema)
        
        cohorte_id = request.args.get('cohorte', type=int)
        if not cohorte_id and cohortes: cohorte_id = cohortes[0]['id']
        
        selected_cohorte = next((c for c in cohortes if c['id'] == cohorte_id), None)
        students, activities, cells = ([], [], {})
        processed_cells = {}
        students_data = {}
        entregas_por_estudiante = {}

        if cohorte_id:
            students, activities, cells = get_matrix_data(conn, schema, cohorte_id)
            now = datetime.now()
            
            for s in students:
                students_data[s['id']] = s
                entregas_por_estudiante[s['id']] = []
                for a in activities:
                    key = (s['id'], a['id'])
                    cell = compute_cell(cells.get(key), a.get('fecha_limite'), now)
                    processed_cells[key] = cell
                    
                    entregas_por_estudiante[s['id']].append({
                        'actividad': a['nombre'],
                        'estado': cell['status'],
                        'color': cell['color'],
                        'tooltip': cell['tooltip']
                    })

        return render_template('index.html', cohortes=cohortes, selected_cohorte=selected_cohorte, 
                               students=students, activities=activities, cells=processed_cells,
                               students_data=students_data, entregas_por_estudiante=entregas_por_estudiante,
                               schema_warnings=_schema_warnings, schema_fixed=_schema_fixed)
    finally:
        conn.close()


@app.route('/resumen')
@requires_auth
def resumen():
    conn = get_db()
    try:
        schema = discover_schema(conn)
        cohortes = get_cohortes(conn, schema)
        cohorte_id = request.args.get('cohorte', type=int)
        if not cohorte_id and cohortes: cohorte_id = cohortes[0]['id']
        selected_cohorte = next((c for c in cohortes if c['id'] == cohorte_id), None)
        
        resumen_data = []
        if cohorte_id:
            students, activities, cells = get_matrix_data(conn, schema, cohorte_id)
            now = datetime.now()
            
            # Identify total integradores and semanales for base calculation
            total_semanales = max(sum(1 for a in activities if a.get('peso', 1) == 1), 1)
            total_integradores = max(sum(1 for a in activities if a.get('peso', 1) > 1), 1)
            
            for s in students:
                semanales_a_tiempo = 0
                semanales_tarde = 0
                integradores_a_tiempo = 0
                integradores_tarde = 0
                tp_integradores_scores = []
                
                for a in activities:
                    is_integrador = a.get('peso', 1) > 1
                    cell = compute_cell(cells.get((s['id'], a['id'])), a.get('fecha_limite'), now)
                    
                    if cell['status'] in ('Entregado', 'Corregido'):
                        if is_integrador: 
                            integradores_a_tiempo += 1
                            tp_integradores_scores.append(100)
                        else: semanales_a_tiempo += 1
                    elif cell['status'] in ('Tarde', 'Rehacer'):
                        if is_integrador: 
                            integradores_tarde += 1
                            tp_integradores_scores.append(50)
                        else: semanales_tarde += 1
                    else:
                        if is_integrador:
                            tp_integradores_scores.append(0)
                        
                participacion_pts = semanales_a_tiempo * 1.0 + semanales_tarde * 0.5
                indice_participacion = min(100, round((participacion_pts / total_semanales) * 100))
                
                tp1 = tp_integradores_scores[0] if len(tp_integradores_scores) > 0 else 0
                tp2 = tp_integradores_scores[1] if len(tp_integradores_scores) > 1 else 0
                
                indice_final = round((tp1 * 0.4) + (tp2 * 0.4) + (indice_participacion * 0.2))
                
                if indice_final >= 90: clasificacion = "Excelente"
                elif indice_final >= 70: clasificacion = "Bueno"
                elif indice_final >= 50: clasificacion = "Regular"
                else: clasificacion = "Deficiente"
                
                resumen_data.append({
                    'id': s['id'],
                    'nombre': s['nombre'],
                    'apellido': s['apellido'],
                    'email': s['email'],
                    'semanales_a_tiempo': semanales_a_tiempo,
                    'semanales_tarde': semanales_tarde,
                    'total_semanales': total_semanales,
                    'porcentaje_semanales': round((semanales_a_tiempo / total_semanales) * 100),
                    'integradores_a_tiempo': integradores_a_tiempo,
                    'total_integradores': total_integradores,
                    'porcentaje_integradores': round((integradores_a_tiempo / total_integradores) * 100),
                    'indice_participacion': indice_participacion,
                    'indice_final': indice_final,
                    'clasificacion': clasificacion
                })
        return render_template('resumen.html', cohortes=cohortes, selected_cohorte=selected_cohorte, resumen=resumen_data)
    finally:
        conn.close()


@app.route('/api/estudiante/<int:estudiante_id>')
def api_estudiante(estudiante_id):
    conn = get_db()
    try:
        schema = discover_schema(conn)
        # Búsqueda simple en la tabla Estudiantes (sin checks rigurosos completos para este API corto)
        table_est = get_table_name(schema, 'Estudiantes')
        col_id = get_column_name(schema, 'Estudiantes', 'Id') or 'id'
        
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM {table_est} WHERE {col_id} = ?", (estudiante_id,))
        estudiante = cursor.fetchone()
        
        if not estudiante:
            return jsonify({"error": "Estudiante no encontrado"}), 404
            
        # Entregas del estudiante
        table_ent = get_table_name(schema, 'Entregas')
        col_ent_est = get_column_name(schema, 'Entregas', 'Estudiante') or 'nc_fk_estudiante_id'
        cursor.execute(f"SELECT * FROM {table_ent} WHERE {col_ent_est} = ?", (estudiante_id,))
        entregas = [dict(row) for row in cursor.fetchall()]
        
        return jsonify({
            "estudiante": dict(estudiante),
            "entregas": entregas
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route('/api/exportar-excel')
@requires_auth
def exportar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    conn = get_db()
    try:
        schema = discover_schema(conn)
        cohortes = get_cohortes(conn, schema)
        cohorte_id = request.args.get('cohorte', type=int)
        if not cohorte_id and cohortes: cohorte_id = cohortes[0]['id']
        selected_cohorte = next((c for c in cohortes if c['id'] == cohorte_id), None)
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Resumen {selected_cohorte['nombre'] if selected_cohorte else 'General'}"
        
        headers = ['ID', 'Apellido', 'Nombre', 'Email', 'Semanales A Tiempo', 'Total Semanales', '% Semanales', 'TP Integradores A Tiempo', 'Total Integradores', '% Integradores', 'Índice Part (%)', 'Índice Final (%)', 'Clasificación']
        ws.append(headers)
        
        # Styles for header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        if cohorte_id:
            students, activities, cells = get_matrix_data(conn, schema, cohorte_id)
            now = datetime.now()
            
            total_semanales = max(sum(1 for a in activities if a.get('peso', 1) == 1), 1)
            total_integradores = max(sum(1 for a in activities if a.get('peso', 1) > 1), 1)
            
            for s in students:
                semanales_a_tiempo = 0
                semanales_tarde = 0
                integradores_a_tiempo = 0
                integradores_tarde = 0
                tp_integradores_scores = []
                
                for a in activities:
                    is_integrador = a.get('peso', 1) > 1
                    cell = compute_cell(cells.get((s['id'], a['id'])), a.get('fecha_limite'), now)
                    
                    if cell['status'] in ('Entregado', 'Corregido'):
                        if is_integrador: 
                            integradores_a_tiempo += 1
                            tp_integradores_scores.append(100)
                        else: semanales_a_tiempo += 1
                    elif cell['status'] in ('Tarde', 'Rehacer'):
                        if is_integrador: 
                            integradores_tarde += 1
                            tp_integradores_scores.append(50)
                        else: semanales_tarde += 1
                    else:
                        if is_integrador: tp_integradores_scores.append(0)
                        
                participacion_pts = semanales_a_tiempo * 1.0 + semanales_tarde * 0.5
                indice_participacion = min(100, round((participacion_pts / total_semanales) * 100))
                
                tp1 = tp_integradores_scores[0] if len(tp_integradores_scores) > 0 else 0
                tp2 = tp_integradores_scores[1] if len(tp_integradores_scores) > 1 else 0
                
                indice_final = round((tp1 * 0.4) + (tp2 * 0.4) + (indice_participacion * 0.2))
                
                if indice_final >= 90: clasificacion = "Excelente"
                elif indice_final >= 70: clasificacion = "Bueno"
                elif indice_final >= 50: clasificacion = "Regular"
                else: clasificacion = "Deficiente"
                
                row = [
                    s['id'], s['apellido'], s['nombre'], s['email'],
                    semanales_a_tiempo, total_semanales, round((semanales_a_tiempo/total_semanales)*100),
                    integradores_a_tiempo, total_integradores, round((integradores_a_tiempo/total_integradores)*100),
                    indice_participacion, indice_final, clasificacion
                ]
                ws.append(row)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter 
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        
        filename = f"resumen_cohorte_{selected_cohorte['nombre']}.xlsx" if selected_cohorte else "resumen.xlsx"
        return send_file(
            out,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error(f"Error generando excel: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()


@app.route('/health')
def health():
    """Endpoint de health-check. Verifica estado de la DB y el esquema."""
    status = {"status": "ok", "tables": {}, "warnings": _schema_warnings, "schema_fixed": _schema_fixed}
    try:
        conn = get_db()
        schema = discover_schema(conn)
        for title in ['Cohortes', 'Actividades', 'Estudiantes', 'Entregas']:
            table_name = get_table_name(schema, title)
            try:
                cursor = conn.cursor()
                cursor.execute(f"SELECT COUNT(*) as cnt FROM {table_name}")
                count = cursor.fetchone()['cnt']
                status['tables'][title] = {"exists": True, "records": count}
            except Exception:
                status['tables'][title] = {"exists": False, "records": 0}
                status['status'] = "degraded"
        conn.close()
    except Exception as e:
        status['status'] = 'error'
        status['error'] = str(e)

    if _schema_warnings:
        status['status'] = 'warning'

    return jsonify(status)


@app.route('/faq')
def faq(): return render_template('faq.html')


# === ARRANQUE ===
# Ejecutar validación al importar el módulo
run_startup_validation()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
